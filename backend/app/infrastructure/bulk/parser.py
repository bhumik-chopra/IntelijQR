import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook
from pydantic import TypeAdapter, ValidationError

from app.core.exceptions import ApplicationError
from app.schemas.qr_generation import QrGenerationRequest


class BulkImportParser:
    _adapter = TypeAdapter(QrGenerationRequest)
    _numeric_fields = {"latitude", "longitude", "max_scans"}
    _boolean_fields = {"hidden"}

    def __init__(self, max_rows: int) -> None:
        self._max_rows = max_rows

    def parse(self, filename: str, content: bytes) -> tuple[list[tuple[int, QrGenerationRequest]], list[dict], int]:
        extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if extension == "csv": rows = self._csv(content)
        elif extension == "xlsx": rows = self._xlsx(content)
        else: raise ApplicationError("BulkForge accepts CSV or XLSX files")
        if not rows: raise ApplicationError("The import file contains no data rows")
        if len(rows) > self._max_rows: raise ApplicationError(f"A bulk job can contain at most {self._max_rows} rows")
        requests: list[tuple[int, QrGenerationRequest]] = []
        errors: list[dict] = []
        for index, row in enumerate(rows, start=2):
            try: requests.append((index, self._adapter.validate_python(self._normalize(row))))
            except (ValidationError, ValueError) as exc: errors.append({"row": index, "message": self._message(exc)})
        return requests, errors, len(rows)

    def _csv(self, content: bytes) -> list[dict]:
        try: text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc: raise ApplicationError("CSV files must use UTF-8 encoding") from exc
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames: raise ApplicationError("The CSV header row is missing")
        rows = []
        for row in reader:
            if any(str(value or "").strip() for value in row.values()): rows.append(dict(row))
            if len(rows) > self._max_rows: break
        return rows

    def _xlsx(self, content: bytes) -> list[dict]:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values)]
            rows = []
            for row in values:
                if any(value not in (None, "") for value in row): rows.append(dict(zip(headers, row)))
                if len(rows) > self._max_rows: break
            workbook.close()
            return rows
        except (StopIteration, ValueError, OSError) as exc: raise ApplicationError("The XLSX workbook could not be read") from exc

    def _normalize(self, row: dict) -> dict:
        values = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
        normalized: dict = {}
        for key, value in values.items():
            if value is None or (isinstance(value, str) and not value.strip()): continue
            value = value.strip() if isinstance(value, str) else value
            if key in self._numeric_fields: value = int(value) if key == "max_scans" else float(value)
            elif key in self._boolean_fields: value = str(value).lower() in {"1", "true", "yes"}
            normalized[key] = value
        if "type" not in normalized: raise ValueError("type is required")
        return normalized

    @staticmethod
    def _message(exc: Exception) -> str:
        if isinstance(exc, ValidationError):
            errors = exc.errors()
            return errors[0].get("msg", "invalid row") if errors else "invalid row"
        return str(exc)
